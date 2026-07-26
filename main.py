from dotenv import load_dotenv
load_dotenv()

import html
import logging
import os
import sqlite3
import time
from contextlib import closing
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup

import dlscore
import stock as stockmod
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

URL = "https://pub-mex.dls.gov.ua/QLA/DocList.aspx"
REQUEST_TIMEOUT = int(os.environ.get("REQUEST_TIMEOUT", "20"))

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "").strip()
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID", "").strip()

DATA_DIR = Path("data")
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / "dls_monitor.db"
XLSX_PATH = DATA_DIR / "Журнал_ДЛС.xlsx"
JSON_PATH = DATA_DIR / "dls.json"

INITIAL_BOOTSTRAP_SILENT = os.environ.get("INITIAL_BOOTSTRAP_SILENT", "true").strip().lower() == "true"

VALID_TYPES = [
    "тимч. заборона",
    "пост. заборона",
    "скасув. тимч. заборони",
    "скасув. пост. заборони",
    "вилучення",
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

JOURNAL_HEADERS = [
    "№ з/п",
    "Дата і номер розпорядження/ рішення/припису",
    "Дата одержання розпорядження/ рішення/припису",
    "Назва лікарських засобів та перелік серій лікарських засобів, зазначених у розпорядженні/ рішенні/приписі",
    "Результати перевірки щодо наявності зазначених лікарських засобів",
    "Вжиті заходи у разі виявлення зазначених лікарських засобів",
    "Дата і номер листа-повідомлення територіальному органу",
    "Підпис уповноваженої особи",
]

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with closing(get_conn()) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS documents (
                uid TEXT PRIMARY KEY,
                doc_num TEXT NOT NULL,
                doc_date TEXT NOT NULL,
                doc_type TEXT NOT NULL,
                drug_name TEXT NOT NULL,
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                telegram_sent INTEGER NOT NULL DEFAULT 0,
                telegram_sent_at TEXT,
                source_seen_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS app_meta (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            """
        )
        # ── Міграція: окремі поля серії та виробника ──────────────────────
        # drug_name склеює назву+серію+виробника і потрібен незмінним, бо
        # з нього будується uid. Ламати uid не можна — бот перешле в Telegram
        # усі 296 записів наново. Тому лише ДОДАЄМО колонки.
        existing = {row[1] for row in conn.execute("PRAGMA table_info(documents)")}
        for col, decl in (
            ("series_raw", "TEXT"),
            ("manufacturer", "TEXT"),
        ):
            if col not in existing:
                conn.execute(f"ALTER TABLE documents ADD COLUMN {col} {decl}")
        conn.commit()


def get_meta(key, default=""):
    with closing(get_conn()) as conn:
        row = conn.execute("SELECT value FROM app_meta WHERE key = ?", (key,)).fetchone()
        return row["value"] if row else default


def set_meta(key, value):
    with closing(get_conn()) as conn:
        conn.execute(
            """
            INSERT INTO app_meta(key, value)
            VALUES(?, ?)
            ON CONFLICT(key) DO UPDATE SET value = excluded.value
            """,
            (key, value),
        )
        conn.commit()


def telegram_enabled():
    return bool(TELEGRAM_TOKEN and TELEGRAM_CHAT_ID)


def build_uid(doc_num, doc_date, drug_name):
    return f"{doc_num.strip()}|{doc_date.strip()}|{drug_name.strip()}".lower()


def get_hidden_fields(soup):
    def val(field_id):
        el = soup.find("input", {"id": field_id})
        return el.get("value", "") if el else ""
    return (
        val("__VIEWSTATE"),
        val("__VIEWSTATEGENERATOR"),
        val("__EVENTVALIDATION"),
    )


def request_with_retry(session, method, url, **kwargs):
    last_error = None
    for attempt in range(1, 4):
        try:
            response = session.request(method, url, timeout=REQUEST_TIMEOUT, **kwargs)
            response.raise_for_status()
            return response
        except Exception as exc:
            last_error = exc
            logging.warning("HTTP %s %s failed (%s/3): %s", method, url, attempt, exc)
            time.sleep(attempt * 2)
    raise last_error


def parse_row(cols):
    type_idx = None
    for idx, col in enumerate(cols):
        text = " ".join(col.stripped_strings).strip().lower()
        if any(v in text for v in VALID_TYPES):
            type_idx = idx
            break
    if type_idx is None:
        return None

    def get(i):
        return " ".join(cols[i].stripped_strings).strip() if i < len(cols) else ""

    return {
        "doc_type": get(type_idx),
        "reg_num": get(type_idx + 1),
        "drug_name": get(type_idx + 2),
        "series": get(type_idx + 3),
        "manufacturer": get(type_idx + 4),
    }


def parse_grid_rows(grid, current_month, current_year):
    records = []
    hit_previous_month = False
    for row in grid.find_all("tr")[1:]:
        cols = row.find_all("td")
        if len(cols) < 3:
            continue

        doc_date = " ".join(cols[0].stripped_strings).strip()
        doc_num = " ".join(cols[1].stripped_strings).strip()
        if not doc_date or not doc_num:
            continue

        try:
            d = datetime.strptime(doc_date, "%d.%m.%Y")
        except ValueError:
            continue

        if d.month != current_month or d.year != current_year:
            hit_previous_month = True
            continue

        parsed = parse_row(cols)
        if not parsed or not parsed["drug_name"]:
            continue

        parts = [parsed["drug_name"]]
        series = parsed["series"]
        if series:
            if "додатку" in series.lower():
                parts.append("(серії зазначені у додатку)")
            else:
                parts.append(f"Серія № {series}")
        if parsed["manufacturer"]:
            parts.append(parsed["manufacturer"])

        drug_full = ", ".join(parts)
        records.append({
            "uid": build_uid(doc_num, doc_date, drug_full),
            "doc_num": doc_num,
            "doc_date": doc_date,
            "doc_type": parsed["doc_type"],
            "drug_name": drug_full,
            # Чисті поля прямо з колонок сайту — саме вони йдуть у dls.json.
            # Розбір склеєного рядка потрібен лише для старих записів.
            "series_raw": series or "",
            "manufacturer": parsed["manufacturer"] or "",
        })
    return records, hit_previous_month


def get_all_documents():
    session = requests.Session()
    now = datetime.now()
    current_month = now.month
    current_year = now.year

    response = request_with_retry(session, "GET", URL, headers=HEADERS)
    soup = BeautifulSoup(response.text, "html.parser")
    vs, vsg, ev = get_hidden_fields(soup)

    payload = {
        "__EVENTTARGET": "ctl00$Content$fvParams$UpdateButton",
        "__EVENTARGUMENT": "",
        "__VIEWSTATE": vs,
        "__VIEWSTATEGENERATOR": vsg,
        "__EVENTVALIDATION": ev,
    }
    response = request_with_retry(session, "POST", URL, headers=HEADERS, data=payload)
    soup = BeautifulSoup(response.text, "html.parser")
    vs, vsg, ev = get_hidden_fields(soup)

    grid = soup.find("table", {"id": "ctl00_Content_gridList"})
    if not grid:
        raise RuntimeError("Таблицю документів не знайдено")

    all_records = []
    page_num = 1

    while True:
        page_records, hit_previous_month = parse_grid_rows(grid, current_month, current_year)
        all_records.extend(page_records)
        logging.info("Сторінка %s: %s записів", page_num, len(page_records))

        if hit_previous_month:
            break

        next_token = f"Page${page_num + 1}"
        has_next = any(next_token in (a.get("href") or "") for a in grid.find_all("a"))
        if not has_next:
            break

        page_num += 1
        payload = {
            "__EVENTTARGET": "ctl00$Content$gridList",
            "__EVENTARGUMENT": f"Page${page_num}",
            "__VIEWSTATE": vs,
            "__VIEWSTATEGENERATOR": vsg,
            "__EVENTVALIDATION": ev,
        }
        response = request_with_retry(session, "POST", URL, headers=HEADERS, data=payload)
        soup = BeautifulSoup(response.text, "html.parser")
        vs, vsg, ev = get_hidden_fields(soup)
        grid = soup.find("table", {"id": "ctl00_Content_gridList"})
        if not grid:
            break

    unique = {}
    for item in all_records:
        unique[item["uid"]] = item
    return list(unique.values())


def send_telegram(doc, stock_block=""):
    if not telegram_enabled():
        logging.warning("Telegram secrets not configured")
        return False

    message = (
        "🚨 <b>Новий припис ДЛС!</b>\n\n"
        f"<b>№</b> {html.escape(doc['doc_num'])}\n"
        f"<b>Дата:</b> {html.escape(doc['doc_date'])}\n"
        f"<b>Захід:</b> {html.escape(doc['doc_type'])}\n"
        f"<b>Ліки:</b> {html.escape(doc['drug_name'])}"
        + stock_block
    )

    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    payload = {
        "chat_id": TELEGRAM_CHAT_ID,
        "text": message,
        "parse_mode": "HTML",
        "disable_web_page_preview": True,
    }

    for attempt in range(1, 4):
        try:
            response = requests.post(url, data=payload, timeout=20)
            if response.status_code == 200:
                logging.info("Telegram sent: %s", doc["doc_num"])
                return True
            logging.warning("Telegram error %s: %s", response.status_code, response.text[:300])
        except Exception as exc:
            logging.warning("Telegram failed (%s/3): %s", attempt, exc)
        time.sleep(attempt * 2)
    return False


def insert_or_update_documents(documents, bootstrap_silent):
    new_count = 0
    telegram_sent_count = 0
    stock_alert_count = 0

    with closing(get_conn()) as conn:
        for doc in documents:
            existing = conn.execute("SELECT uid FROM documents WHERE uid = ?", (doc["uid"],)).fetchone()
            if existing:
                # Доливаємо чисті поля і в старі рядки, де їх ще немає
                conn.execute(
                    """
                    UPDATE documents
                       SET source_seen_at = CURRENT_TIMESTAMP,
                           series_raw   = COALESCE(NULLIF(?, ''), series_raw),
                           manufacturer = COALESCE(NULLIF(?, ''), manufacturer)
                     WHERE uid = ?
                    """,
                    (doc.get("series_raw") or "", doc.get("manufacturer") or "", doc["uid"]),
                )
                continue

            telegram_sent = 1 if bootstrap_silent else 0
            telegram_sent_at = datetime.utcnow().isoformat(timespec="seconds") if bootstrap_silent else None

            conn.execute(
                """
                INSERT INTO documents(uid, doc_num, doc_date, doc_type, drug_name,
                                      series_raw, manufacturer, telegram_sent, telegram_sent_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (doc["uid"], doc["doc_num"], doc["doc_date"], doc["doc_type"], doc["drug_name"],
                 doc.get("series_raw"), doc.get("manufacturer"), telegram_sent, telegram_sent_at),
            )
            new_count += 1
        conn.commit()

    if not bootstrap_silent:
        with closing(get_conn()) as conn:
            unsent = conn.execute(
                """
                SELECT uid, doc_num, doc_date, doc_type, drug_name, series_raw, manufacturer
                FROM documents
                WHERE telegram_sent = 0
                ORDER BY substr(doc_date, 7, 4) || '-' || substr(doc_date, 4, 2) || '-' || substr(doc_date, 1, 2), doc_num, drug_name
                """
            ).fetchall()

        # Залишки тягнемо один раз і лише якщо є що розсилати. Якщо Google Drive
        # недоступний — stock_items = None, повідомлення все одно піде, просто
        # без блоку про склад. Втратити алерт через збій Drive неприпустимо.
        stock_items = stockmod.fetch_stock() if unsent else None
        stock_index = stockmod.build_index(stock_items) if stock_items else {}

        for doc in unsent:
            doc = dict(doc)
            try:
                rec = dlscore.build_record(
                    doc["uid"], doc["doc_num"], doc["doc_date"], doc["doc_type"],
                    doc["drug_name"],
                    series_raw=doc.get("series_raw") or None,
                    manufacturer=doc.get("manufacturer") or None,
                )
                hits = stockmod.find_hits(rec, stock_items, stock_index)
                stock_block = stockmod.format_hits(hits)
                if hits:
                    stock_alert_count += 1
            except Exception as exc:
                logging.warning("Звірка із залишками не вдалась для %s: %s", doc["doc_num"], exc)
                stock_block = ""

            if send_telegram(doc, stock_block):
                with closing(get_conn()) as conn:
                    conn.execute(
                        "UPDATE documents SET telegram_sent = 1, telegram_sent_at = ? WHERE uid = ?",
                        (datetime.utcnow().isoformat(timespec="seconds"), doc["uid"]),
                    )
                    conn.commit()
                telegram_sent_count += 1

    return new_count, telegram_sent_count, stock_alert_count


MONTHS_UA = [
    "Січень", "Лютий", "Березень", "Квітень", "Травень", "Червень",
    "Липень", "Серпень", "Вересень", "Жовтень", "Листопад", "Грудень",
]


def month_title(ym):
    """'2026-04' → 'Квітень 2026'"""
    try:
        y, m = ym.split("-")
        return f"{MONTHS_UA[int(m) - 1]} {y}"
    except Exception:
        return ym


def load_records():
    """Читає БД і будує нормалізовані записи з обчисленим актуальним статусом."""
    with closing(get_conn()) as conn:
        rows = conn.execute(
            "SELECT uid, doc_num, doc_date, doc_type, drug_name, series_raw, manufacturer FROM documents"
        ).fetchall()

    records = []
    for r in rows:
        r = dict(r)
        records.append(
            dlscore.build_record(
                r["uid"], r["doc_num"], r["doc_date"], r["doc_type"], r["drug_name"],
                series_raw=r.get("series_raw") or None,
                manufacturer=r.get("manufacturer") or None,
            )
        )
    return dlscore.compute_status(records)


def export_json():
    """
    data/dls.json — те, що читає Liki.on!.
    Віддається через raw.githubusercontent.com (CORS: *), тож фронт бере
    його напряму, без проксі й без Pages.
    """
    records = load_records()
    payload = dlscore.build_payload(
        records, datetime.utcnow().isoformat(timespec="seconds") + "Z"
    )
    JSON_PATH.write_text(dlscore.dumps(payload), encoding="utf-8")
    return payload


def export_excel(records=None):
    """
    Журнал ДЛС: окремий аркуш на кожен місяць, нумерація з початку місяця,
    лише чинні заборони (скасовані не потрапляють).

    Колонка 5 тут завжди «відсутній» — цей файл не знає про залишки складів.
    Заповнену версію генерує Liki.on! у браузері, де залишки вже є.
    """
    if records is None:
        records = load_records()

    active = [r for r in records if r["active"]]

    by_month = {}
    for r in active:
        by_month.setdefault(r["ym"], []).append(r)

    font = Font(name="Verdana", size=8)
    header_font = Font(name="Verdana", size=8, bold=True)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    widths = [6, 30, 20, 60, 30, 30, 30, 20]

    wb = Workbook()
    wb.remove(wb.active)

    for ym in sorted(by_month.keys()):
        ws = wb.create_sheet(title=month_title(ym)[:31])
        ws.append(JOURNAL_HEADERS)
        ws.append(["1", "2", "3", "4", "5", "6", "7", "8"])

        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        rows = sorted(by_month[ym], key=lambda r: (r["date"][6:], r["date"][3:5], r["date"][:2], r["num"]))
        for index, r in enumerate(rows, start=1):
            full = r.get("full") or ", ".join(
                x for x in [
                    r["name"],
                    ("Серія № всі серії" if r["all_series"] else
                     ("Серія № " + ", ".join(r["series"]) if r["series"] else "")),
                    r.get("manufacturer") or "",
                ] if x
            )
            ws.append([
                index,
                f"№ {r['num']} від {r['date']}",
                r["date"],
                full,
                "відсутній",
                "", "", "",
            ])

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=8):
            for cell in row:
                cell.font = header_font if cell.row == 1 else font
                cell.alignment = align_center
                cell.border = border
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 40

    if not wb.sheetnames:
        wb.create_sheet(title="Журнал ДЛС")

    wb.save(XLSX_PATH)


def main():
    init_db()
    documents = get_all_documents()
    bootstrap_done = get_meta("bootstrap_done", "false").lower() == "true"
    bootstrap_silent = (not bootstrap_done) and INITIAL_BOOTSTRAP_SILENT

    new_documents, telegram_sent_count, stock_alerts = insert_or_update_documents(documents, bootstrap_silent)

    records = load_records()
    payload = export_json()
    export_excel(records)

    if not bootstrap_done:
        set_meta("bootstrap_done", "true")

    print(f"documents_found={len(documents)}")
    print(f"new_documents={new_documents}")
    print(f"telegram_sent={telegram_sent_count}")
    print(f"stock_alerts={stock_alerts}")
    print(f"db_path={DB_PATH}")
    print(f"xlsx_path={XLSX_PATH}")
    print(f"json_path={JSON_PATH}")
    print(f"active_bans={payload['active_bans']}")
    print(f"parse_failures={sum(1 for r in payload['records'] if not r['parse_ok'])}")


if __name__ == "__main__":
    main()
