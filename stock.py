"""
Читання залишків складів ТМО і звірка їх із приписами ДЛС — серверна частина.

Потрібно тільки для того, щоб Telegram-повідомлення про новий припис одразу
казало, чи є ця серія на складах. Liki.on! робить таку саму звірку в браузері;
логіка навмисно продубльована один-в-один.

УВАГА: колонки й правила відсіву заголовків мають збігатися з parseWorkbook()
та isDeptHeader() в index.html. Якщо там зміниться мапінг — міняти і тут.
"""

import io
import logging
import os

import requests
from openpyxl import load_workbook

import dlscore

# Той самий файл, який читає «Пошук медикаментів» у Liki.on!
STOCK_URL = os.environ.get(
    "STOCK_XLSX_URL",
    "https://docs.google.com/spreadsheets/d/1uAByiAlSotEn1Y7fO1_fN72vOZ8lcu1z/export?format=xlsx",
)
STOCK_TIMEOUT = int(os.environ.get("STOCK_TIMEOUT", "60"))

# A(0)=Артикул, E(4)=МНН, M(12)=Номенклатура, O(14)=Серія,
# R(17)=Од.вим., S(18)=Поч.залишок, V(21)=Кінц.залишок
COL_A, COL_E, COL_M, COL_O, COL_R, COL_S, COL_V = 0, 4, 12, 14, 17, 18, 21

SKIP_HEADERS = [
    "артикул", "параметри", "відомість", "кількість товарів", "сегодня",
    "номенклатура", "характеристика", "серія", "придатний", "відбір",
    "разом", "од. вим",
]
SKIP_EXACT_HEADERS = ["склад"]


def _cell(row, idx):
    if idx >= len(row):
        return ""
    v = row[idx]
    return "" if v is None else str(v).strip()


def _is_dept_header(row):
    a = _cell(row, COL_A)
    m = _cell(row, COL_M)
    if not a or len(a) < 3:
        return False
    if m:                      # рядки з даними завжди мають назву в M
        return False
    if a.isdigit():            # артикул «202»
        return False
    lc = a.lower()
    if lc in SKIP_EXACT_HEADERS:
        return False
    if any(lc.startswith(k) for k in SKIP_HEADERS):
        return False
    return True


def parse_stock(data):
    """bytes xlsx → [{dept, name, mnn, series, unit, qty}]"""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    items = []
    dept = "—"
    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        if _is_dept_header(row):
            dept = _cell(row, COL_A)
            continue
        name = _cell(row, COL_M)
        if not name or name.lower() == "номенклатура":
            continue
        raw_qty = row[COL_V] if len(row) > COL_V and row[COL_V] not in (None, "") else (
            row[COL_S] if len(row) > COL_S else None
        )
        try:
            qty = float(raw_qty) if raw_qty not in (None, "") else None
        except (TypeError, ValueError):
            qty = None
        items.append({
            "dept": dept,
            "name": name,
            "mnn": _cell(row, COL_E) or None,
            "series": _cell(row, COL_O) or None,
            "unit": _cell(row, COL_R) or None,
            "qty": qty,
        })
    wb.close()
    return items


def fetch_stock():
    """Повертає список позицій або None, якщо не вдалось. Ніколи не кидає."""
    try:
        resp = requests.get(STOCK_URL, timeout=STOCK_TIMEOUT)
        resp.raise_for_status()
        items = parse_stock(resp.content)
        logging.info("Залишки: %s позицій", len(items))
        return items
    except Exception as exc:
        logging.warning("Не вдалось прочитати залишки складів: %s", exc)
        return None


def build_index(items):
    """згорнута серія → [позиції]"""
    idx = {}
    for it in items:
        if not it["series"]:
            continue
        k = dlscore.fold(it["series"])
        if k:
            idx.setdefault(k, []).append(it)
    return idx


def _hay(it):
    return dlscore.fold_name(it["name"]) + " " + dlscore.fold_name(it.get("mnn") or "")


def find_hits(record, items, index):
    """
    Шукає позиції складу, що підпадають під припис.
    Повертає [{dept, series, qty, unit}] згорнуте по відділенню+серії.
    """
    if items is None:
        return None
    hits = []

    if record["all_series"]:
        # Серії немає — матчимо за назвою, двома словами, щоб не ловити
        # хибні збіги на кшталт «СПИРТ»
        toks = _brand_tokens(record["brand"], wide=True)
        if toks:
            for it in items:
                h = _hay(it)
                if all(t in h for t in toks):
                    hits.append((it, it["series"] or "—"))
    else:
        for s in record["series"]:
            for it in index.get(dlscore.fold(s), []):
                hits.append((it, s))

    agg = {}
    for it, series in hits:
        key = (it["dept"], series)
        cur = agg.setdefault(key, {"dept": it["dept"], "series": series,
                                   "qty": 0.0, "unit": it["unit"] or ""})
        cur["qty"] += (it["qty"] or 0)
        if not cur["unit"] and it["unit"]:
            cur["unit"] = it["unit"]
    return sorted(agg.values(), key=lambda v: (v["dept"], v["series"]))


def _brand_tokens(brand, wide=False):
    words = [w for w in dlscore.fold_name(brand).split(" ") if len(w) >= 4]
    if not words:
        return []
    return words[:2] if wide else words[:1]


def format_hits(hits):
    """Блок для Telegram-повідомлення."""
    if hits is None:
        return "\n\n<i>⚠️ Залишки складів перевірити не вдалось</i>"
    if not hits:
        return "\n\n✅ <b>На залишках ТМО відсутній</b>"
    lines = ["", "", "🚨 <b>Є НА СКЛАДАХ ТМО:</b>"]
    for h in hits[:15]:
        qty = int(h["qty"]) if float(h["qty"]).is_integer() else round(h["qty"], 2)
        lines.append(f"• {h['dept']} — сер. {h['series']}, <b>{qty} {h['unit']}</b>".rstrip())
    if len(hits) > 15:
        lines.append(f"• …та ще {len(hits) - 15}")
    lines.append("")
    lines.append("<i>Перевірити та вилучити з обігу.</i>")
    return "\n".join(lines)
